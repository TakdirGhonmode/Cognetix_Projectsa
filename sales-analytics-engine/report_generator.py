import os
import logging
from typing import Dict, Any, List
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

class ReportGenerator:
    """
    Export & Reporting module.
    Handles CSV exports, cleaned dataset exports, multi-tab openpyxl Excel reports, 
    and optional HTML dashboard generation.
    """

    def __init__(self, output_dir: str = "output"):
        self.output_dir = output_dir
        os.makedirs(self.output_dir, exist_ok=True)

    def export_csv_reports(
        self,
        cleaned_df: pd.DataFrame,
        kpi_summary: pd.DataFrame,
        monthly_df: pd.DataFrame,
        regional_df: pd.DataFrame
    ) -> Dict[str, str]:
        """
        Exports CSV summary reports and cleaned dataset.
        """
        paths = {}

        cleaned_path = os.path.join(self.output_dir, "cleaned_sales_data.csv")
        cleaned_df.to_csv(cleaned_path, index=False)
        paths["cleaned_sales_data"] = cleaned_path

        kpi_path = os.path.join(self.output_dir, "kpi_summary.csv")
        kpi_summary.to_csv(kpi_path, index=False)
        paths["kpi_summary"] = kpi_path

        monthly_path = os.path.join(self.output_dir, "monthly_trends.csv")
        monthly_df.to_csv(monthly_path, index=False)
        paths["monthly_trends"] = monthly_path

        regional_path = os.path.join(self.output_dir, "regional_performance.csv")
        regional_df.to_csv(regional_path, index=False)
        paths["regional_performance"] = regional_path

        logger.info(f"CSV reports exported to {self.output_dir}")
        return paths

    def export_excel_report(
        self,
        cleaned_df: pd.DataFrame,
        kpis: Dict[str, Any],
        insights: List[str],
        regional_df: pd.DataFrame,
        product_df: pd.DataFrame,
        monthly_df: pd.DataFrame,
        file_name: str = "sales_report.xlsx"
    ) -> str:
        """
        Generates a management-ready, formatted Excel report using openpyxl.
        Tabs: Executive Summary, Cleaned Data, Regional & Segment Stats, Monthly Trends.
        """
        file_path = os.path.join(self.output_dir, file_name)
        wb = openpyxl.Workbook()
        
        # Define Styling Theme
        header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        title_font = Font(name="Calibri", size=16, bold=True, color="1F4E78")
        card_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        card_font = Font(name="Calibri", size=11, bold=True, color="000000")
        thin_border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )

        # -------------------------------------------------------------
        # Tab 1: Executive Summary
        # -------------------------------------------------------------
        ws_exec = wb.active
        ws_exec.title = "Executive Summary"
        ws_exec.views.sheetView[0].showGridLines = True

        ws_exec["A1"] = "SALES PERFORMANCE ANALYTICS — EXECUTIVE SUMMARY"
        ws_exec["A1"].font = title_font

        # KPI Metric Cards
        ws_exec["A3"] = "Key Performance Indicator (KPI)"
        ws_exec["B3"] = "Value"
        ws_exec["A3"].font = header_font
        ws_exec["A3"].fill = header_fill
        ws_exec["B3"].font = header_font
        ws_exec["B3"].fill = header_fill

        kpi_items = [
            ("Total Sales Revenue", f"${kpis['Total_Sales_Revenue']:,.2f}"),
            ("Total Units Sold", f"{kpis['Total_Units_Sold']:,}"),
            ("Total Transactions", f"{kpis['Total_Transactions']:,}"),
            ("Average Order Value (AOV)", f"${kpis['Average_Order_Value']:,.2f}"),
            ("Average Unit Price", f"${kpis['Average_Unit_Price']:,.2f}")
        ]

        row_idx = 4
        for metric, val in kpi_items:
            ws_exec.cell(row=row_idx, column=1, value=metric).font = card_font
            ws_exec.cell(row=row_idx, column=1).fill = card_fill
            ws_exec.cell(row=row_idx, column=2, value=val).font = card_font
            ws_exec.cell(row=row_idx, column=2).fill = card_fill
            ws_exec.cell(row=row_idx, column=1).border = thin_border
            ws_exec.cell(row=row_idx, column=2).border = thin_border
            row_idx += 1

        # Key Business Insights Section
        row_idx += 2
        ws_exec.cell(row=row_idx, column=1, value="Strategic Business Insights & Recommendations").font = title_font
        row_idx += 1
        for ins in insights:
            ws_exec.cell(row=row_idx, column=1, value=f"• {ins}").font = Font(name="Calibri", size=10)
            row_idx += 1

        # -------------------------------------------------------------
        # Tab 2: Monthly Trends
        # -------------------------------------------------------------
        ws_monthly = wb.create_sheet(title="Monthly Trends")
        ws_monthly.views.sheetView[0].showGridLines = True
        self._write_df_to_sheet(ws_monthly, monthly_df, header_fill, header_font, thin_border)

        # -------------------------------------------------------------
        # Tab 3: Regional & Product Stats
        # -------------------------------------------------------------
        ws_stats = wb.create_sheet(title="Regional & Product Stats")
        ws_stats.views.sheetView[0].showGridLines = True
        
        ws_stats["A1"] = "Regional Sales Performance"
        ws_stats["A1"].font = title_font
        self._write_df_to_sheet(ws_stats, regional_df, header_fill, header_font, thin_border, start_row=3)

        start_prod_row = len(regional_df) + 6
        ws_stats.cell(row=start_prod_row, column=1, value="Product Performance Breakdown").font = title_font
        self._write_df_to_sheet(ws_stats, product_df, header_fill, header_font, thin_border, start_row=start_prod_row + 2)

        # -------------------------------------------------------------
        # Tab 4: Cleaned Sales Data
        # -------------------------------------------------------------
        ws_data = wb.create_sheet(title="Cleaned Data")
        ws_data.views.sheetView[0].showGridLines = True
        # Export subset of clean data to avoid bloat if large
        export_df = cleaned_df.copy()
        if "Date" in export_df.columns:
            export_df["Date"] = export_df["Date"].dt.strftime("%Y-%m-%d")
        self._write_df_to_sheet(ws_data, export_df, header_fill, header_font, thin_border)

        # Auto-adjust column widths across all sheets
        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_len = 0
                col_letter = openpyxl.utils.get_column_letter(col[0].column)
                for cell in col:
                    if cell.value:
                        val_str = str(cell.value)
                        if len(val_str) > max_len:
                            max_len = len(val_str)
                sheet.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 60)

        wb.save(file_path)
        logger.info(f"Management-ready Excel report exported to: {file_path}")
        return file_path

    def _write_df_to_sheet(self, ws, df: pd.DataFrame, header_fill, header_font, thin_border, start_row: int = 1):
        """Helper to write DataFrame to openpyxl sheet with header styling."""
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=start_row):
            for c_idx, val in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=val)
                cell.border = thin_border
                if r_idx == start_row:
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center", vertical="center")

    def export_optional_html_report(
        self,
        kpis: Dict[str, Any],
        insights: List[str],
        file_name: str = "sales_report.html"
    ) -> str:
        """
        OPTIONAL: Generates a responsive visual HTML Executive Dashboard.
        """
        file_path = os.path.join(self.output_dir, file_name)
        
        insight_html = "".join([f"<li>{ins}</li>" for ins in insights])

        html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Sales Performance Analytics Dashboard</title>
    <style>
        body {{ font-family: 'Segoe UI', Tahoma, Geneva, Verdana, sans-serif; background: #0f172a; color: #f8fafc; margin: 0; padding: 30px; }}
        .header {{ text-align: center; margin-bottom: 30px; }}
        .header h1 {{ font-size: 2.2rem; color: #38bdf8; margin-bottom: 5px; }}
        .kpi-container {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(200px, 1fr)); gap: 20px; margin-bottom: 30px; }}
        .kpi-card {{ background: #1e293b; border-radius: 12px; padding: 20px; text-align: center; border: 1px solid #334155; }}
        .kpi-title {{ font-size: 0.9rem; color: #94a3b8; margin-bottom: 8px; font-weight: 600; }}
        .kpi-value {{ font-size: 1.6rem; color: #f1f5f9; font-weight: bold; }}
        .section {{ background: #1e293b; padding: 25px; border-radius: 12px; border: 1px solid #334155; margin-bottom: 30px; }}
        .section h2 {{ color: #38bdf8; margin-top: 0; }}
        ul {{ line-height: 1.8; color: #cbd5e1; }}
        .charts-grid {{ display: grid; grid-template-columns: repeat(auto-fit, minmax(400px, 1fr)); gap: 20px; }}
        .chart-card {{ background: #1e293b; border-radius: 12px; padding: 15px; border: 1px solid #334155; text-align: center; }}
        .chart-card img {{ max-width: 100%; border-radius: 8px; }}
    </style>
</head>
<body>
    <div class="header">
        <h1>Sales Performance Analytics Engine</h1>
        <p style="color: #94a3b8;">Management Business Intelligence Executive Dashboard</p>
    </div>

    <div class="kpi-container">
        <div class="kpi-card">
            <div class="kpi-title">TOTAL REVENUE</div>
            <div class="kpi-value">${kpis['Total_Sales_Revenue']:,.2f}</div>
        </div>
        <div class="kpi-card">
            <div class="kpi-title">UNITS SOLD</div>
            <div class="kpi-value">{kpis['Total_Units_Sold']:,}</div>
        </div>
        <div class="kpi-card">
            <div class="kpi-title">TOTAL TRANSACTIONS</div>
            <div class="kpi-value">{kpis['Total_Transactions']:,}</div>
        </div>
        <div class="kpi-card">
            <div class="kpi-title">AVERAGE ORDER VALUE</div>
            <div class="kpi-value">${kpis['Average_Order_Value']:,.2f}</div>
        </div>
    </div>

    <div class="section">
        <h2>Executive Business Insights & Recommendations</h2>
        <ul>
            {insight_html}
        </ul>
    </div>

    <div class="charts-grid">
        <div class="chart-card">
            <h3>Monthly Sales Trend</h3>
            <img src="charts/monthly_sales_trend.png" alt="Monthly Sales Trend">
        </div>
        <div class="chart-card">
            <h3>Regional Performance</h3>
            <img src="charts/regional_performance.png" alt="Regional Performance">
        </div>
        <div class="chart-card">
            <h3>Revenue by Category</h3>
            <img src="charts/revenue_by_category.png" alt="Revenue by Category">
        </div>
        <div class="chart-card">
            <h3>Customer Segment Matrix</h3>
            <img src="charts/customer_segment_matrix.png" alt="Customer Segment Matrix">
        </div>
    </div>
</body>
</html>
"""
        with open(file_path, "w", encoding="utf-8") as f:
            f.write(html_content)
        
        logger.info(f"Optional HTML Dashboard exported to: {file_path}")
        return file_path
